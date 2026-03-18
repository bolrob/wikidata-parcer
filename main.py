import requests
import time
from openpyxl import load_workbook
from datetime import datetime
import re

def query_wikidata(sparql_query, limit=1000, offset=0):
    url = "https://query.wikidata.org/sparql"
    headers = {
        "User-Agent": "MyBot/1.0 (your-email@example.com)",
        "Accept": "application/json"
    }

    if "LIMIT" not in sparql_query.upper():
        sparql_query += f" LIMIT {limit} OFFSET {offset}"

    params = {
        "query": sparql_query,
        "format": "json"
    }
    response = requests.get(url, params=params, headers=headers, timeout=30)

    if response.status_code == 200:
        data = response.json()
        return data["results"]["bindings"]
    else:
        print(f"Ошибка {response.status_code}: {response.text[:200]}")
        return []

def get_participants(war_qid):
    query = f"""
    SELECT ?participantLabel WHERE {{
      wd:{war_qid} wdt:P710 ?participant .
      SERVICE wikibase:label {{ bd:serviceParam wikibase:language "en". }}
    }}
    """
    results = query_wikidata(query, limit=500)
    participants = [res["participantLabel"]["value"] for res in results if "participantLabel" in res]
    return participants

def parse_wikidata_date(date_str):
    if not date_str:
        return None
    match = re.search(r'[+-]?(\d{4})-(\d{2})-(\d{2})', date_str)
    if match:
        year, month, day = map(int, match.groups())
        return datetime(year, month, day)
    return None

def date_to_year(dt):
    return dt.year if dt else None

def war_in_range(war_start, war_end, range_start, range_end):
    if war_start is None and war_end is None:
        return False

    if war_start is not None and war_end is None:
        return war_start <= range_end

    if war_start is None and war_end is not None:
        return war_end >= range_start

    return war_start <= range_end and war_end >= range_start



start_time = time.perf_counter()

try:
    start_year = int(input("Введите начальный год диапазона: "))
    end_year = int(input("Введите конечный год диапазона: "))
    if start_year > end_year:
        start_year, end_year = end_year, start_year
except ValueError:
    print("Некорректный ввод. Будет использован диапазон 1800-2100.")
    start_year, end_year = 1800, 2100

range_start = datetime(start_year, 1, 1)
range_end = datetime(end_year, 12, 31)

class_qid = 'Q198'
base_query = f"""
SELECT DISTINCT ?item ?itemLabel ?start ?end ?point WHERE {{
  ?item wdt:P31/wdt:P279* wd:{class_qid} .
  OPTIONAL {{ ?item wdt:P580 ?start . }}
  OPTIONAL {{ ?item wdt:P582 ?end . }}
  OPTIONAL {{ ?item wdt:P585 ?point . }}
  SERVICE wikibase:label {{ bd:serviceParam wikibase:language "en". }}
}}
"""

all_wars = []
offset = 0
limit = 100
while True:
    print(f"Загружаем войны, offset={offset}...")
    batch = query_wikidata(base_query, limit=limit, offset=offset)
    if not batch:
        break
    all_wars.extend(batch)
    offset += limit
    #time.sleep(1)

print(f"Всего загружено войн: {len(all_wars)}")

fn = "example.xlsx"
try:
    wb = load_workbook(fn)
    ws = wb['Лист1']
except FileNotFoundError:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Лист1'
    headers = ['id', 'Название', 'Начало', 'Конец', 'url', 'Участники']
    for col, val in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = val
else:
    if ws.max_column < 6:
        ws.cell(row=1, column=6).value = 'Участники'

row = ws.max_row + 1

for war in all_wars:
    item_url = war["item"]["value"]
    item_qid = item_url.split("/")[-1]
    label = war.get("itemLabel", {}).get("value", "нет метки")

    start_date_str = war.get("start", {}).get("value", None)
    end_date_str = war.get("end", {}).get("value", None)
    point_str = war.get("point", {}).get("value", None)

    start_dt = parse_wikidata_date(start_date_str)
    end_dt = parse_wikidata_date(end_date_str)
    point_dt = parse_wikidata_date(point_str)

    if start_dt is None and point_dt is not None:
        start_dt = point_dt
        start_date_str = point_str

    if not war_in_range(start_dt, end_dt, range_start, range_end):
        continue

    participants = get_participants(item_qid)
    participants_str = ", ".join(participants) if participants else "нет данных"
    #time.sleep(0.5)

    ws.cell(row=row, column=1).value = item_qid
    ws.cell(row=row, column=2).value = label
    ws.cell(row=row, column=3).value = start_date_str if start_date_str else "нет данных"
    ws.cell(row=row, column=4).value = end_date_str if end_date_str else "нет данных"
    ws.cell(row=row, column=5).value = item_url
    ws.cell(row=row, column=6).value = participants_str

    row += 1

wb.save(fn)
wb.close()

end_time = time.perf_counter()
print(f"Добавлено записей: {row - 2}")
print(f"Время выполнения: {end_time - start_time:.6f} секунд")